# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


OUTPUT = "User_Stories_Sport_Cloud_360.xlsx"


BACKLOG_HEADERS = [
    "US_ID",
    "Epic",
    "Persona",
    "User Story",
    "Critères d'acceptation",
    "Objets Salesforce impactés",
    "Permission Set",
    "Automatisation",
    "Priorité",
    "Sprint cible",
    "Statut",
]

BACKLOG_ROWS = [
    [
        "US-001",
        "Core Platform",
        "Admin Sport Cloud",
        "En tant qu'admin, je veux configurer les effectifs.",
        "CRUD Squad__c + tri DisplayOrder__c",
        "Squad__c",
        "SportCloud Admin",
        "Aucune",
        "P1",
        "Sprint 1",
        "To Do",
    ],
    [
        "US-002",
        "Athlete 360",
        "Coach",
        "En tant que coach, je veux consulter la fiche 360 d'un joueur.",
        "Vue profil/performance/médical/administratif",
        "Athlete__c",
        "Coach Read",
        "Aucune",
        "P1",
        "Sprint 1",
        "To Do",
    ],
    [
        "US-003",
        "Athlete 360",
        "Admin effectif",
        "En tant qu'admin, je veux historiser les changements d'effectif.",
        "Affectation datée valide, pas de fin < début",
        "SquadAssignment__c",
        "Admin Effectif",
        "Flow/Apex",
        "P1",
        "Sprint 1",
        "To Do",
    ],
    [
        "US-004",
        "Scouting 360",
        "Scout",
        "En tant que scout, je veux créer une piste de recrutement.",
        "Prospect créé avec statut initial",
        "Prospect__c",
        "Scout Editor",
        "Aucune",
        "P1",
        "Sprint 1",
        "To Do",
    ],
    [
        "US-005",
        "Scouting 360",
        "Responsable recrutement",
        "En tant que recruteur, je veux convertir un prospect en joueur.",
        "Création Athlete__c + lien ConvertedAthlete__c",
        "Prospect__c, Athlete__c",
        "Recruiter + Direction",
        "Flow/Apex",
        "P1",
        "Sprint 2",
        "To Do",
    ],
    [
        "US-006",
        "Planning",
        "Staff",
        "En tant que staff, je veux voir le planning dans le calendrier standard.",
        "Événements visibles dans Event/Calendar",
        "PlanningEvent__c, Event",
        "Coach/Staff Read",
        "Apex Trigger",
        "P1",
        "Sprint 1",
        "Done",
    ],
    [
        "US-007",
        "Planning",
        "Logistique",
        "En tant que logistique, je veux réserver une ressource sans conflit.",
        "Réservation refusée en cas de chevauchement",
        "Booking__c, Resource__c",
        "Logistics Editor",
        "Flow/Apex",
        "P1",
        "Sprint 2",
        "To Do",
    ],
    [
        "US-008",
        "Scolarité",
        "Référent academy",
        "En tant que référent, je veux suivre la scolarité des jeunes.",
        "Dossier annuel + bulletins + absences",
        "AthleteEducation__c, EducationReport__c, SchoolAbsence__c",
        "Academy Editor",
        "Flow",
        "P1",
        "Sprint 2",
        "To Do",
    ],
    [
        "US-009",
        "Data",
        "Data Analyst",
        "En tant qu'analyste, je veux centraliser les données externes.",
        "Datapoints liés à athlete/prospect + source",
        "DataSource__c, ExternalAthleteData__c, ExternalProspectData__c",
        "Data Analyst",
        "Apex (ingestion)",
        "P1",
        "Sprint 2",
        "To Do",
    ],
    [
        "US-010",
        "Sécurité",
        "Compliance",
        "En tant que compliance, je veux restreindre l'accès médical.",
        "Accès médical réservé aux profils habilités",
        "MedicalEvent__c, RecoveryPlan__c",
        "Medical Only",
        "Sharing Rules",
        "P1",
        "Sprint 1",
        "To Do",
    ],
]

PERM_HEADERS = [
    "Rôle",
    "Module",
    "Objet",
    "Lire",
    "Créer",
    "Modifier",
    "Supprimer",
    "Restrictions de champs",
    "Commentaire",
]

PERM_ROWS = [
    ["SportCloud Admin", "All", "*", "Oui", "Oui", "Oui", "Oui", "Aucune", "Administration complète"],
    ["Coach", "Athlete/Planning", "Athlete__c", "Oui", "Non", "Limité", "Non", "Pas de détails médicaux sensibles", "Consultation joueur"],
    ["Medical Staff", "Médical", "MedicalEvent__c", "Oui", "Oui", "Oui", "Non", "Champs médicaux stricts", "Suivi santé"],
    ["Scout", "Scouting", "Prospect__c", "Oui", "Oui", "Oui", "Non", "Pas de données médicales", "Pistes recrutement"],
    ["Academy Staff", "Scolarité", "AthleteEducation__c", "Oui", "Oui", "Oui", "Non", "Jeunes uniquement", "Suivi scolaire"],
]

AUTO_HEADERS = [
    "Automation_ID",
    "Déclencheur",
    "Règle métier",
    "Type",
    "Résultat attendu",
    "Gestion d'erreur",
    "Owner",
    "Priorité",
]

AUTO_ROWS = [
    [
        "AUTO-001",
        "PlanningEvent__c (after insert/update/delete/undelete)",
        "Synchroniser vers Event calendrier",
        "Apex Trigger",
        "Event créé/mis à jour/supprimé",
        "Logs + reprise manuelle",
        "Admin Planning",
        "P1",
    ],
    [
        "AUTO-002",
        "SquadAssignment__c (create/update)",
        "Mettre à jour CurrentSquad__c",
        "Flow/Apex",
        "Effectif courant fiable",
        "Validation anti-chevauchement",
        "Admin Effectif",
        "P1",
    ],
    [
        "AUTO-003",
        "Prospect__c (status=Converted)",
        "Créer Athlete et lier ConvertedAthlete__c",
        "Flow",
        "Conversion scouting -> athlete",
        "Rollback + message utilisateur",
        "Recruiter",
        "P1",
    ],
]

SPRINT_HEADERS = ["Sprint", "US_ID", "Effort", "Dépendances", "Risque", "Owner", "Date cible"]

SPRINT_ROWS = [
    ["Sprint 1", "US-001", "S", "Aucune", "Faible", "Admin", "2026-03-02"],
    ["Sprint 1", "US-002", "M", "US-001", "Faible", "Coach Lead", "2026-03-03"],
    ["Sprint 1", "US-006", "M", "US-001", "Moyen", "Planning Lead", "2026-03-04"],
    ["Sprint 2", "US-005", "L", "US-004", "Moyen", "Recruiting Lead", "2026-03-11"],
    ["Sprint 2", "US-008", "M", "US-001", "Faible", "Academy Lead", "2026-03-10"],
]


def apply_format(ws, headers, priority_column_name=None):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    priority_fills = {
        "P1": PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid"),
        "P2": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "P3": PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid"),
    }

    for idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    for idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 65)

    if priority_column_name and priority_column_name in headers:
        pcol = headers.index(priority_column_name) + 1
        for r in range(2, ws.max_row + 1):
            p = ws.cell(row=r, column=pcol).value
            if p in priority_fills:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = priority_fills[p]


def write_sheet(ws, headers, rows, priority_column_name=None):
    ws.append(headers)
    for row in rows:
        ws.append(row)
    apply_format(ws, headers, priority_column_name)


def main():
    wb = Workbook()
    ws_backlog = wb.active
    ws_backlog.title = "Backlog"
    ws_perm = wb.create_sheet("Permissions_Matrix")
    ws_auto = wb.create_sheet("Automations_Matrix")
    ws_sprint = wb.create_sheet("Sprint_Plan")

    write_sheet(ws_backlog, BACKLOG_HEADERS, BACKLOG_ROWS, "Priorité")
    write_sheet(ws_perm, PERM_HEADERS, PERM_ROWS)
    write_sheet(ws_auto, AUTO_HEADERS, AUTO_ROWS, "Priorité")
    write_sheet(ws_sprint, SPRINT_HEADERS, SPRINT_ROWS)

    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
